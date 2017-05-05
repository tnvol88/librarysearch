from googleplaces import GooglePlaces, types, lang
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, colors


API_KEY = 'AIzaSyDLdDH5kqx4-kYy3qlTpxjGc0KEVD9orN4' #Enter your key between quotes
google_places = GooglePlaces(API_KEY)

wb = load_workbook('library.xlsx')
ws = wb.active


"""
These variables are used to define fill colors
for changing cell colors based on library search
results
"""
redFill = PatternFill('solid',fgColor='FF0000') #Found one exact match and is closed
greenFill = PatternFill('solid',fgColor='00FF00') #Found one exact match and is open
yellowFill = PatternFill('solid',fgColor='FFFF00') #Name not exact, address not exact
cyanFill = PatternFill('solid',fgColor='00FFFF') #Name not exact, address exact, is open
magentaFill = PatternFill('solid',fgColor='FF00FF') #Name not exact, address exact, is closed
blueFill = PatternFill('solid',fgColor='0000FF') # Found zero results

not_found_libraries = {}
closed_dict = {}
open_dict = {}
same_address = {}
same_number = {}

def checkIfClosed(library_name, row):
    """
    Searches for a library and then finds whether library
    is permanently closed, is open, or is not found.
    Input: library name
    Returns: 'not found': if search returns nothing
            'False': if library is permanently closed
            'True': if library is still open
    """
    row = row
    print('Checking row ', row)
    query_result = google_places.nearby_search(
        location=str(ws.cell(row=row, column=9).value)+','+str(ws.cell(row=row,column=11).value),
        keyword=str(library_name)+' '+str(ws.cell(row=row, column=9).value),
        rankby='distance')
    if len(query_result.places)== 0:
        not_found_libraries[library_name]={'found':'Zero Results', 'address from Google':'Zero Results',
                                           'address from Excel':str(ws.cell(row=row, column=7).value.strip()),
                                           'row':row
                                           }
        return 'no match'
    else:
        place, match = findBestResult(query_result.places, library_name, row)
        if place == None:
            not_found_libraries[library_name]={'found':'No matches', 'address from Google':'No matches',
                                           'address from Excel':str(ws.cell(row=row, column=7).value.strip()),
                                               'row':row
                                           }
            return None, None
        place.get_details()
        address_from_web = addressChange(place.formatted_address)
        address_from_excel = str(ws.cell(row=row, column =7).value).strip()
        try:
            _ = (place.details['permanently_closed'])
            closed_dict[library_name]={'found':place.name, 'address from Google':place.formatted_address,
                                      'address from Excel':address_from_excel
                                      }
            return 'Closed', match
        except KeyError:
            open_dict[library_name]={'found':place.name, 'address from Google':place.formatted_address,
                                     'address from Excel':address_from_excel
                                     }
            return 'Open', match

def findBestResult(results, library_name, row):
    """Iterates through each result found. First checks if result matches name of library in excel,
        if not a match, it then tests whether address is the same. If neither test returns a correct
        match, no result is returned.
    """
    row = row
    for place in results:
        address_from_excel = str(ws.cell(row=row, column =7).value).strip()
        place.get_details()
        googlenumber = place.local_phone_number
        excelnumber = ws.cell(row=row, column=13).value
        numberCheck(excelnumber, googlenumber)
        if str(place.name) == str(library_name):
            return place, None
        elif addressCheck(addressChange(place.formatted_address), row):
            print('address check: True')
            same_address[library_name]={'found':place.name, 'address from Google':place.formatted_address,
                                        'address from Excel': address_from_excel, 'row':row
                                        }
            return place, 'notexact'
        elif numberCheck(excelnumber, googlenumber):
            same_number[library_name]={'found':place.name, 'address from Google':place.formatted_address,
                                       'address from Excel': address_from_excel, 'row':row
                                       }
            return place, 'notexact'
        else:
            pass
    return None, None

def addressCheck(address, row):
    """This function removes numbers from address
       and returns True if street matches and False
       if they do not match.
    """
    address_from_excel = str(ws.cell(row=row, column =7).value).strip()
    address_from_excel = addressChange(address_from_excel)
    print('address from excel: ', address_from_excel)
    print('address from web: ', address)
    return address == address_from_excel

def addressChange(address):
    abbreviations = {'st':'street', 'rd':'road', 'ave':'avenue',
                     'dr':'drive', 'n':'north', 's':'south',
                     'e':'east', 'w':'west', 'ln':'lane', 'cres':'cresent'
                     }
    address = address.lower()
    try:
        address = address[:address.index(',')]
    except ValueError:
        pass
    address = address.split(' ')
    if address[0].isdigit():
        address = address[1:]
    other = address[:]
    for i in address:
        if i in abbreviations:
            other[address.index(i)] = abbreviations[i]
    return other
    
def changenumber(number):
    correctednumber = []
    for i in number:
        if i.isdigit():
            correctednumber.append(i)
    return ''.join(correctednumber)

def numberCheck(excelnumber, googlenumber):
    excelnumber = changenumber(excelnumber)
    googlenumber = changenumber(googlenumber)
    if googlenumber in excelnumber:
        print('numbers match')
        return True
    else:
        print('numbers do not match')
        return False
"""
The loop below loops through values in column F (col 6)
and runs the above function using that value. Depending on the result
of the function run, the cell color is changed.
"""
for i in range(500,550): # change '40' to '4041' to run script on entire column
    col = 6
    library = (ws.cell(row=i,column=col).value)
    values = checkIfClosed(library,i)
    if values == 'no match':
        ws.cell(row=i, column=col).fill=blueFill
    if values[0]=='Open' and values[1]==None:
        ws.cell(row=i,column=col).fill=greenFill
    elif values[0] == 'Closed' and values[1]==None:
        ws.cell(row=i,column=col).fill=redFill
    elif values[0] == None and values[1] == None:
        ws.cell(row=i, column=col).fill = yellowFill
    elif values[0] == 'Open' and values[1] == 'notexact':
        ws.cell(row=i, column=col).fill = cyanFill
    elif values[0] == 'Closed' and values[1] == 'notexact':
        ws.cell(row=i, column=col).fill = magentaFill

wb.save('library.xlsx') # saves changes to excel workbook



"""
Following print statements print a list of
libraries which were not found, permanently closed, or
still open. This may be useful for double checking closed
libraries or not found libraries. The values which are printed
in these list reflect the text Google used to search.
"""

print('\n')
print('*'*20)
print('For the following, Google found a place with a different name but on the same street in the same town')
print('*'*20, '\n')
for k in same_address:
    print('searched for: ', k)
    print('row: ', same_address[k]['row'])
    print('found: ', same_address[k]['found'])
    print('address from Google: ', same_address[k]['address from Google'])
    print('address from Excel: ', same_address[k]['address from Excel'])
    print('\n')

print('*'*20)
print('For the following, no results were found matching name or address')
print('*'*20, '\n')
for k in not_found_libraries:
    print('searched for: ', k)
    print('row: ', not_found_libraries[k]['row'])
    print('found: ', not_found_libraries[k]['found'])
    print('address from Google: ', not_found_libraries[k]['address from Google'])
    print('address from Excel: ', not_found_libraries[k]['address from Excel'])
    print('\n')

print('*'*20)
print('For the following, Google found a place with different name and address, but with the same phone number')
print('*'*20)
for k in same_number:
    print('searched for: ', k)
    print('row: ', same_number[k]['row'])
    print('found: ', same_number[k]['found'])
    print('address from Google: ', same_number[k]['address from Google'])
    print('address from Excel: ', same_number[k]['address from Excel'])
    print('\n')


